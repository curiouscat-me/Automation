import pandas as pd
from config.config_GUI import *
from modules.logger import data_logger
import os
from openpyxl import load_workbook
from identifier_functions_FTE.identifier_FTE_GUI import *
from modules.formatting import *
from modules.eofy import get_eofy
from modules.get_column_index import get_column_index
from modules.date_extraction import extract_date_from_filename
from modules.skip_column import initiate_skip_column_fte
from modules.missing_employees import identify_missing_employees_fte
from modules.check_mark_fulfilled import check_and_mark_fulfilled_fte
from modules.sanity_checks import sanity_checks_fte
from modules.kill_switch import terminate_process
 
"""
Global Parameters:
    current_df (DataFrame): DataFrame containing the current month data from the Static Report.
    next_df (DataFrame): DataFrame containing the next month data from the Static Report.
    op_fte_df (DataFrame): DataFrame containing the operational plan data.
    original_op_fte_df (DataFrame): The original DataFrame.
    ws (Worksheet): The worksheet to apply the highlights.
    output_directory (str): The directory where the output file will be saved.
"""

def load_data():
    """
    Loads data from the operational plan and static files based on configuration.
    
    Returns:
    tuple: DataFrames for current month, next month, operational plan, and original operational plan.
    
    Process:
    1. Define a mapping dictionary for employee group names.
    2. Load current month and next month data from the two Static Report sheets.
    3. Apply employee group mapping and rename columns based on the configuration.
    4. Filter data for the Security domain entries.
    5. Log the count of records loaded from the Static Report.
    6. Load the Op Plan FTE data.
    7. Filter out rows with 'Vacant', 'Role Handed Back', missing 'LANID', or 'FTE Resource Type' in the 'Resource Type' column.
    8. Remove duplicates based on 'Employee ID'.
    9. Log the count of records in the Op Plan FTE sheet.
    10. Create a copy of the Op Plan for comparison later.
    11. Return the loaded DataFrames.
    If an error occurs during the process, logs the error and returns None for all DataFrames.
    """
    try:
        """
        Loads data from the operational plan and static files based on configuration.
        Returns DataFrames for current month, next month, and operational plan.
        """

        # Define a mapping dictionary for employee group names
        employee_group_mapping = {
            'Permanent Employee': 'Permanent',
            'Fixed Term Employee': 'Fixed Term Contract',
        }

        data_logger.info(f"Loading Static Report data from {CONFIG['STATIC_FILE']}...")
        # Load current month and next month data from the two Static Report sheets.
        current_df = pd.read_excel(CONFIG['STATIC_FILE'], sheet_name=CONFIG['sheets']['current_month']['sheet_name'],
                                   usecols=CONFIG['sheets']['current_month']['usecols'], header=1)
        next_df = pd.read_excel(CONFIG['STATIC_FILE'], sheet_name=CONFIG['sheets']['next_month']['sheet_name'],
                                usecols=CONFIG['sheets']['next_month']['usecols'], header=1) 

        # Pre-formatting for Employee Category
        for df in [current_df, next_df]:
            df['Employee Group (Name)'] = df['Employee Group (Name)'].replace(employee_group_mapping) # Apply employee mapping for FTE and MS
            df.rename(columns=CONFIG['COLUMN_MAPPING_FTE'], inplace=True)
        data_logger.info("Employee Category mapping has been applied for Static Report.")

        # Filter for Security domain entries
        current_security_count = len(current_df[(current_df['Domain'] == 'Security') & (current_df['FTE Category'] == 'FTE')])
        next_security_count = len(next_df[(next_df['Domain'] == 'Security') & (next_df['FTE Category'] == 'FTE')])
        
        data_logger.info(f"Security domain: {current_security_count} records from current month, {next_security_count} records from next month.")

        # Load the Op Plan data
        data_logger.info(f"Loading Op Plan FTE data from {CONFIG['OP_FILE']}...")
        op_fte_df = pd.read_excel(CONFIG['OP_FILE'], sheet_name=CONFIG['sheets']['FTE']['sheet_name'],
                                  usecols=CONFIG['sheets']['FTE']['usecols'], header=3)
        
        # Filter out rows with 'Vacant' in the 'FTE Name' column
        op_fte_df = op_fte_df[~op_fte_df['FTE Name'].str.contains('Vacant', na=False) |
                              ~op_fte_df['FTE Name'].str.contains('Role Handed Back', na=False) |
                              ~op_fte_df['LANID'].isna() |
                              ~op_fte_df['Resource Type'].str.contains('FTE Resource Type', na=False)]

        # Removing duplicates based on Employee ID
        unique_op_fte_df = op_fte_df.drop_duplicates(subset=['Employee ID'])
        
        op_fte_count = len(unique_op_fte_df[(unique_op_fte_df['Resource Type'] != "FTE Resource Type") & 
                                            (unique_op_fte_df['LANID'] != "") &
                                            (unique_op_fte_df['FTE Name'] != "Role Handed Back") &
                                            (unique_op_fte_df['Employee ID'].notna())])

        data_logger.info(f"Security domain: {op_fte_count} unique entries from FTE sheet.")
        data_logger.info("The figures stated above are estimates, as these also accounted for duplicated entries and cancelled/vacants, etc.")

        # Create a copy of Op Plan for Comparision later
        original_op_fte_df = op_fte_df.copy()

        return current_df, next_df, op_fte_df, original_op_fte_df

    except Exception as e:
        data_logger.error(f"Error loading data: {e}")
        return None, None, None, None

def merge_data_fte(current_df, next_df, op_fte_df):
    """
    Merges the operational plan data with the current month and next month data
    from the static report. Set datetime object. Initializes the 'Modified' and 'Line Manager' columns.
    
    Returns:
    DataFrame: Merged DataFrame containing all the data with initialized columns.
    
    Process:
    1. Log the start of data processing.
    2. Retrieve and log the current End of Financial Year (EOFY).
    3. Get and log the current date.
    4. Extract and check the file date from the static report filename.
    5. Rename columns in current_df and next_df to match the Op Plan columns based on the provided mapping.
    6. Initialize the 'Modified' column in op_fte_df.
    7. Initialize the 'Line Manager' column in op_fte_df.
    8. Log the completion of data merging and column initialization.
    9. Return the updated op_fte_df.
    """
    data_logger.info("Starting data processing...")
    eofy = get_eofy()
    data_logger.info(f"The current End of Financial Year (EOFY) is: {eofy.date()}")

    current_date = datetime.now().strftime('%d-%m-%y')
    data_logger.info(f"The current date is: {current_date}")

    # Extract and check the file date
    file_date = extract_date_from_filename(CONFIG['STATIC_FILE'])
    if not file_date:
        data_logger.error("Failed to extract date from Static Report")
        return None

    # Rename columns in current_df and next_df to match the Op Plan columns based on the mapping
    current_df = current_df.rename(columns=CONFIG['COLUMN_MAPPING_FTE'])
    next_df = next_df.rename(columns=CONFIG['COLUMN_MAPPING_FTE'])

    # Initialize 'Modified' column
    op_fte_df['Modified'] = ""  # Mark all entries as initially unmodified

    # Initialize 'Line Manager' column
    op_fte_df['Line Manager'] = ""  # Mark all entries as initially unmodified

    data_logger.info("Data merging and columns initiated.")
    return op_fte_df

def highlight_differences(ws, op_fte_df, original_op_fte_df):
    """
    Highlights cells in the modified DataFrame that are different from the original DataFrame.
    Stops checking at the row where multiple columns contain 'x'.
    Skips multiple columns as there are not enough sufficient information for comparison.
    
    Process:
    1. Define the fill pattern for modified cells.
    2. Extract headers and create mappings for original and processed data indices.
    3. Determine the row to stop highlighting based on 'x' count in columns A to G.
    4. Normalize values for comparison.
    5. Identify columns to skip based on predefined skip list.
    6. Iterate over each row in the worksheet:
       a. Skip rows based on specific conditions (e.g., 'Vacant', 'Role handed back').
       b. For matching rows in original and processed data, compare cell values.
       c. Highlight cells with differences using the defined fill pattern.
    """
    modified_fill = PatternFill(start_color="7EC8E3", end_color="7EC8E3", fill_type="solid")

    headers = {cell.value: cell.column_letter for cell in ws[1]}
    original_index_map = original_op_fte_df.groupby(['Employee ID', 'FTE Name', 'Start Date']).apply(lambda x: x.index.tolist()).to_dict()
    processed_index_map = op_fte_df.groupby(['Employee ID', 'FTE Name', 'Start Date']).apply(lambda x: x.index.tolist()).to_dict()

    stop_highlighting_row = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        x_count = sum(cell.value == 'x' for cell in row)
        if x_count > 1:
            stop_highlighting_row = row[0].row
            break
    
    skip_columns_list = [
        'Input Annualised Stretch $ \n(if applicable)', 'Squad', 'Service', 'Asset', 'Product', 
        'Tech Financial Reform', 'FTET approval Ref', 'FTE #', 'Headcount', 'Start Date', 'Comments',
        'Free Input', 'Run %', 'Divisional Change %', 'Tech Projects %', 'Total %'
    ]

    # Identify columns to skip based on header names
    skip_columns = [col_letter for col_name, col_letter in headers.items() if col_name in skip_columns_list]

    for row in ws.iter_rows(min_row=2, max_row=(stop_highlighting_row or ws.max_row) - 1):
        row_idx = row[0].row - 2
        resource_type = op_fte_df.at[row_idx, 'Resource Type']
        employee_id = op_fte_df.at[row_idx, 'Employee ID']
        fte_name = op_fte_df.at[row_idx, 'FTE Name']
        lanid = op_fte_df.at[row_idx, 'LANID']
        # tech_area = op_fte_df.at[row_idx, 'Tech Area']
        start_date = op_fte_df.at[row_idx, 'Start Date']

        if 'Vacant' in str(fte_name) or 'Role handed back' in str(fte_name) or pd.isna(lanid) or pd.isna(employee_id) or resource_type == 'Stretch':
            continue
        
        if (employee_id, fte_name, start_date) in original_index_map:
            original_row_indices = original_index_map[employee_id, fte_name, start_date]
            processed_row_indices = processed_index_map[employee_id, fte_name, start_date]
            for original_row_idx in original_row_indices:
                for processed_row_idx in processed_row_indices:
                    for col_name, col_letter in headers.items():
                        if col_name in original_op_fte_df.columns and col_name in op_fte_df.columns:
                            if col_letter in skip_columns:
                                continue
                            
                            original_value = original_op_fte_df.at[original_row_idx, col_name]
                            modified_value = op_fte_df.at[processed_row_idx, col_name]
                            
                            if pd.isna(modified_value) or pd.isna(original_value):
                                continue

                            original_value = normalize(original_value)
                            modified_value = normalize(modified_value)

                            if original_value != modified_value:
                                cell = row[ws[col_letter + str(row[0].row)].col_idx - 1]
                                cell.fill = modified_fill

                                # Debug print to see what's being compared
                                print(f"Highlighting cell {cell.coordinate} - Original: {original_value}, Modified: {modified_value}")

def highlight_vacant_stretch(ws, op_fte_df):
    """
    Highlights cells in the worksheet based on specific conditions:
    - Vacant positions are highlighted in yellow.
    - Stretch roles are highlighted in black.
    
    Process:
    1. Define fill patterns for vacant and stretch roles.
    2. Iterate over each row in the worksheet from the second row to the last, specifically in column D.
    3. Retrieve the 'FTE Name' and 'Resource Type' for the corresponding row from op_fte_df.
    4. For each cell in the row:
       a. If 'FTE Name' contains 'Vacant', apply the vacant_fill pattern.
       b. If 'Resource Type' contains 'Stretch', apply the stretch_fill pattern.
    """
    vacant_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    stretch_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=4, max_col=4):
        row_idx = row[0].row - 2
        fte_name = op_fte_df.at[row_idx, 'FTE Name']
        resource_type = op_fte_df.at[row_idx, 'Resource Type']
        for cell in row:
            if 'Vacant' in str(fte_name):
                cell.fill = vacant_fill
            elif 'Stretch' in str(resource_type):
                cell.fill = stretch_fill

def save_data(op_fte_df, original_op_fte_df, output_directory):
    """
    Saves the processed data to an Excel file with specific formatting applied to rows based on their role status.
    Includes rows with specific keywords from the original dataset.
    
    Returns:
    str: The path to the saved Excel file.
    
    Process:
    1. Generate a timestamp and create the output file path.
    2. Check if the output file already exists and log if it will be overwritten.
    3. Filter out rows where "Resource Type" is "FTE Resource Type".
    4. Save the filtered DataFrame to an Excel file.
    5. Load the workbook and worksheet from the saved file.
    6. Get column indices for 'LANID', 'FTE Name', 'Start Date', and 'End Date'.
    7. Format duplicated 'LANID' cells.
    8. Apply date formatting to 'Start Date' and 'End Date' columns.
    9. Highlight differences between the processed and original DataFrames.
    10. Highlight vacant and stretch roles.
    11. Save the formatted workbook.
    12. Log the save operation and return the output file path.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d")
    output_file = os.path.join(output_directory, f"Security (FTE) - {timestamp}.xlsx")

    # Check if the output file already exists
    if os.path.exists(output_file):
        data_logger.info(f"File already exists. Overwriting... {output_file}")

    # Filter out rows where "Resource Type" = "FTE Resource Type"
    filtered_df = op_fte_df[(op_fte_df['Resource Type'] != "FTE Resource Type")]
    filtered_df.to_excel(output_file, index=False, sheet_name='FTE', engine='openpyxl')

    # Load workbook and worksheet
    wb = load_workbook(output_file)
    ws = wb.active

    lanid_column_index = get_column_index(ws, 'LANID')
    fte_name_column_index = get_column_index(ws, 'FTE Name')
    format_duplicate_lanid_fte(ws, lanid_column_index, fte_name_column_index)

    start_date_index = get_column_index(ws, 'Start Date')
    end_date_index = get_column_index(ws, 'End Date')

    # Apply comprehensive formatting to the worksheet
    apply_date_format(wb, ws, start_date_index, end_date_index)

    # Highlight modified cells
    highlight_differences(ws, op_fte_df, original_op_fte_df)
    highlight_vacant_stretch(ws, op_fte_df)

    wb.save(output_file)
    data_logger.info(f"Data saved to {output_file}")

    return output_file

def process_fte(op_plan_path, static_report_path, output_directory, result):
    """
    Executes the FTE Op Plan Automation Script.
    
    Parameters:
    op_plan_path (str): The file path to the operational plan.
    static_report_path (str): The file path to the static report.
    output_directory (str): The directory where the output file will be saved.
    result: Required if you are running this from the GUI, if you want to run it directly from fte_GUI.py, 
    then remove this parameter and provide the path to file at the end of the script.
    
    Process:
    1. Set the file paths for the operational plan and static report.
    2. Log the start of the process and the current timestamp.
    3. Load data from the specified files.
       - If data loading fails, log an error and terminate the script.
    4. Check for termination signal after each major step.
    5. Identify missing employees from the original operational plan.
    6. Mark fulfilled rows.
    7. Initialize the 'Skip' column.
    8. Perform sanity checks on the data.
    9. Extract the date from the static report filename for consistent use.
       - If date extraction fails, log an error and terminate the script.
    10. Merge the operational plan data with current and next month data.
        - If merging fails, log an error and terminate the script.
    11. Process data through various scenario functions to identify specific changes:
        - Exits, new joiners, transfers in/out, grade changes, internal mobility, conversions, line manager changes, location changes.
    12. Save the processed data to an Excel file.
    13. Highlight differences and vacant/stretch roles in the saved workbook.
    14. Save the workbook after highlighting differences.
    15. Log the completion time of the script and the duration of the execution.
    
    Exceptions:
    - Logs any unexpected errors and terminates the process gracefully.
    """
    set_op_plan_path(op_plan_path)
    set_static_report_path(static_report_path)

    try:
        data_logger.info('-----------------------------------------------------------------------------------------------------------------------------------')
        data_logger.info("Executing FTE Op Plan Automation Script...")
        start_time = datetime.now()

        # Load Data
        current_df, next_df, op_fte_df, original_op_fte_df = load_data()
        if op_fte_df is None:
            data_logger.error("Failed to load Op Plan FTE data. Exiting script")
            return  # Ensure proper termination without using sys.exit()

        # Check for termination
        if terminate_process.is_set():
            data_logger.info("Process terminated by user.")
            return

        # Identify missing employees from the original Op Plan
        op_fte_df = identify_missing_employees_fte(current_df, next_df, op_fte_df)

        # Check for termination
        if terminate_process.is_set():
            data_logger.info("Process terminated by user.")
            return

        # Mark fulfilled rows
        op_fte_df = check_and_mark_fulfilled_fte(current_df, next_df, op_fte_df)

        # Check for termination
        if terminate_process.is_set():
            data_logger.info("Process terminated by user.")
            return

        # Initiate Skip column
        op_fte_df = initiate_skip_column_fte(op_fte_df)

        # Check for termination
        if terminate_process.is_set():
            data_logger.info("Process terminated by user.")
            return

        # Perform sanity checks
        op_fte_df = sanity_checks_fte(current_df, next_df, op_fte_df)

        # Extract the date once and use it throughout to avoid repeated extraction
        file_date = extract_date_from_filename(CONFIG['STATIC_FILE'])
        if not file_date:
            data_logger.error("Failed to extract date from Static Report. Exiting script")
            return

        # Process data (initial processing)
        op_fte_df = merge_data_fte(current_df, next_df, op_fte_df)
        if op_fte_df is None:
            data_logger.error("Merging data process failed. Exiting script")
            return
        
        # Process data through various scenario functions
        scenario_functions_fte = [
            identify_exits_fte, # TODO - Finished
            identify_new_joiners_fte, # TODO - Finished
            identify_transfers_in_fte, # TODO - Finished
            identify_transfers_out_fte, # TODO - Finished
            identify_grade_changes_fte, # TODO - Finished
            identify_internal_mobility_fte, # TODO: Need fix multiple roles
            indetify_conversions_within_fte, # TODO - Finished
            identify_conversions_cwr_to_fte, # TODO - Finished
            identify_conversions_fte_to_cwr, # TODO - Finished
            identify_line_manager_changes_fte, # TODO - Finished
            identify_location_changes_fte # TODO - Finished
        ]

        for func in scenario_functions_fte:
            # Check for termination
            if terminate_process.is_set():
                data_logger.info("Process terminated by user.")
                return
            op_fte_df = func(current_df, next_df, op_fte_df, file_date)

        # Save Data
        output_file = save_data(op_fte_df, original_op_fte_df, output_directory)

        # Highlight differences
        wb = load_workbook(output_file)
        ws = wb.active
        highlight_differences(ws, op_fte_df, original_op_fte_df)
        highlight_vacant_stretch(ws, op_fte_df)

        # Save the workbook after highlighting differences
        wb.save(output_file)

        # Log script completion time
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        data_logger.info(f"Script execution completed successfully in {duration} seconds.")
    
    except Exception as e:
        data_logger.error(f"An unexpected error occurred, terminating process: {e}.", exc_info=True)

"""
Provide the path to the file you want to run the script directly from the file.

if __name__ == "__main__":
    test_op_plan_path = "My Documents\\An\\VS Learning\\Practice\\New-Test files\\Copy Op SEC May.xlsx"
    test_static_report_path = "My Documents\\An\VS Learning\\Practice\\New-Test files\\Copy of Static 240630.xlsx"
    test_output_directory = "SHARE\\App Support\\Support\\3. Teams\\Technology Business Operations\\4. Practice Management\\Automation Tool\\Test Folder"
    process_fte(test_op_plan_path, test_static_report_path, test_output_directory)
"""