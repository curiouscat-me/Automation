import re
import pandas as pd
from datetime import datetime
from openpyxl.styles import NamedStyle
from modules.get_column_index import get_column_index
from modules.logger import data_logger
from openpyxl.styles import PatternFill, NamedStyle
import collections

def format_tech_area(name):
    """
    Formats the tech area name to ensure it adheres to specific naming conventions.
    
    Parameters:
    name (str): The tech area name to format.
    Returns:
    str: The formatted tech area name. If the input is not a string, returns the original value.
    
    Process:
    1. Check if the input name is a string.
    2. If the name does not start with "Security_", prepend "Security_" to the name.
    3. Replace spaces, commas, plus signs, hyphens, and parentheses with underscores.
    4. Remove leading and trailing underscores.
    5. Return the formatted name.
    6. If the input is not a string, return the original value.
    """
    if isinstance(name, str):
        # Replace spaces, commas, plus signs, hyphens, and parentheses with underscores
        name = re.sub(r'[\s,+\-()]+', '_', name).strip('_')
        
        # Ensure the name starts with "Security_" and format appropriately
        if not name.startswith("Security_"):
            name = "Security_" + name

        # Replace spaces, commas, plus signs, hyphens, and parentheses with underscores
        name = re.sub(r'[\s,+\-()]+', '_', name).strip('_')
        return name
    else:
        # Return the original value if it's not a string
        return name

def format_domain(name):
    """
    Formats the domain name to ensure it adheres to specific naming conventions.
    
    Parameters:
    name (str): The tech area name to
    Returns:
    str: The formatted domain name. If the input is not a string, returns the original value.
    
    Process:
    1. Check if the input name is a string.
    2. If the name does not end with "Domain", append "_Domain" to the name.
    3. Replace spaces and commas with underscores.
    4. Remove leading and trailing underscores.
    5. Return the formatted name.
    6. If the input is not a string, return the original value.
    """
    if isinstance(name, str):
        # Ensure the name ends with "Domain" and format appropriately
        if not name.endswith("Domain"):
            name = name + "_Domain"

        # Replace spaces and commas with a single underscore
        name = re.sub(r'[\s,-]+', '_', name).strip('_')
        return name
    else:
        return name
 
def map_to_hub_FTE(country):
    """
    Maps a given country to its corresponding hub for FTE.
    
    Parameters:
    country (str): The name of the country to map.
    
    Returns:
    str: The corresponding hub name if the country is in the mapping; otherwise, returns the original country name.
    
    Process:
    1. Define a dictionary mapping specific countries to their corresponding hubs.
    2. Use the dictionary to map the given country to its hub.
    3. If the country is not found in the dictionary, return the original country name.
    """
    hub_mapping = {
        'India': 'India Hub',
        'Philippines': 'Manila Hub',
        'Australia': 'Australia',
        # Add more if required
    }
    return hub_mapping.get(country, country)

def apply_date_format(wb, ws, start_date_index, end_date_index):
    """
    Applies a custom date format to the 'Start Date' and 'End Date' columns in the given worksheet.
    
    Parameters:
    wb (Workbook): The workbook containing the worksheet.
    ws (Worksheet): The worksheet to apply the date formatting.
    start_date_index (int): The index of the 'Start Date' column.
    end_date_index (int): The index of the 'End Date' column.
    
    Process:
    1. Locate the indices for the 'Start Date' and 'End Date' columns.
    2. Define a custom date format style (MMM-YY).
    3. Add the custom date style to the workbook.
    4. Apply the custom date format to the cells in the 'Start Date' and 'End Date' columns for all rows starting from the second row.
    5. Verify that the columns are found.
       - If any of the columns are missing, log an error.
       - If both columns are found, log a success message.
    """
    # Locate necessary column indices
    start_date_index = get_column_index(ws, 'Start Date')
    end_date_index = get_column_index(ws, 'End Date')

    # Define date format 
    date_style = NamedStyle(name='custom_datetime', number_format='MMM-YY')
    wb.add_named_style(date_style)

    # Apply date format to Start Date and End Date columns
    for row in ws.iter_rows(min_col=start_date_index, max_col=end_date_index, min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.style = date_style

    # Verify that the columns are found
    if start_date_index is None or end_date_index is None:
        data_logger.error("One or more necessary date columns are missing")
        return
    data_logger.info("Start Date and End Date formatting has been applied successfully!")

def format_duplicate_lanid_fte(ws, lanid_column_index, fte_name_column_index):
    """
    Highlights cells in the LANID column that have duplicate values with a specific fill color.
    
    Parameters:
    ws (Worksheet): The worksheet to apply the highlights.
    lanid_column_index (int): The index of the LANID column.
    fte_name_column_index (int): The index of the FTE Name column.
    
    Process:
    1. Define a fill pattern for duplicate LANID cells (light red).
    2. Get all the cells in the LANID column from the worksheet.
    3. Extract the values from the LANID cells, skipping blanks.
    4. Identify duplicate LANIDs.
    5. Find the index for the 'FTE Name' column.
    6. Iterate over the rows in the worksheet:
       a. Skip rows with blank LANID values.
       b. Retrieve the FTE Name value, handling null or blank values explicitly.
       c. If the LANID is in the set of duplicates, apply the fill pattern to the LANID cell.
    """
    duplicate_fill = PatternFill(start_color='FFADB0', end_color='FFADB0', fill_type='solid') # Light Red for duplicated LANID
    # Get all the cells in the LANID column as a list of tuples
    lanid_cells = list(ws.iter_cols(min_col=lanid_column_index, max_col=lanid_column_index, min_row=2, max_row=ws.max_row))[0]
    # Extract the values from the cells
    lanids = [cell.value for cell in lanid_cells if cell.value]
    # Find duplicates by seeing which LANID appears more than once
    duplicate_lanids = {lanid for lanid, count in collections.Counter(lanids).items() if count > 1}

    # Find index for 'FTE Name' column
    fte_name_column_index = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == 'FTE Name':
            fte_name_column_index = idx + 1
            break

    # Iterate over the rows and apply the fill if the LANID is in the set of duplicates
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        lanid_cell = row[lanid_column_index - 1] # Adjust for 0-based index
                # Skip rows with blank LANID
        if not lanid_cell.value:
            continue
        
        fte_name_cell = row[fte_name_column_index - 1] # -1 to convert back to 0-based indexing
        # Handle null or blank values explicitly
        fte_name_value = fte_name_cell.value or ""
        fte_name_value = fte_name_value.lower().strip()

        # Apply cell format to LANID cell if it's a duplicate
        if lanid_cell.value in duplicate_lanids: # lanid_cell is specifically targeted within each row
            lanid_cell.fill = duplicate_fill


def format_duplicate_lanid_ms(ws, lanid_column_index, ms_name_column_index):
    """
    Highlights cells in the LANID column that have duplicate values with a specific fill color.
    
    Parameters:
    ws (Worksheet): The worksheet to apply the highlights.
    lanid_column_index (int): The index of the LANID column.
    ms_name_column_index (int): The index of the Resource Name column.
    
    Process:
    1. Define a fill pattern for duplicate LANID cells (light red).
    2. Get all the cells in the LANID column from the worksheet.
    3. Extract the values from the LANID cells, skipping blanks.
    4. Identify duplicate LANIDs.
    5. Find the index for the 'Resource Name' column.
    6. Iterate over the rows in the worksheet:
       a. Skip rows with blank LANID values.
       b. Retrieve the Resource Name value, handling null or blank values explicitly.
       c. If the LANID is in the set of duplicates, apply the fill pattern to the LANID cell.
    """
    duplicate_fill = PatternFill(start_color='FFADB0', end_color='FFADB0', fill_type='solid') # Light Red for duplicated LANID
    # Get all the cells in the LANID column as a list of tuples
    lanid_cells = list(ws.iter_cols(min_col=lanid_column_index, max_col=lanid_column_index, min_row=2, max_row=ws.max_row))[0]
    # Extract the values from the cells
    lanids = [cell.value for cell in lanid_cells]
    # Find duplicates by seeing which LANID appears more than once
    duplicate_lanids = {lanid for lanid, count in collections.Counter(lanids).items() if count > 1}

    # Find index for 'Resource Name' column
    ms_name_column_index = None
    for idx, cell in enumerate(ws[1]):
        if cell.value == 'Resource Name':
            ms_name_column_index = idx + 1
            break

    # Iterate over the rows and apply the fill if the LANID is in the set of duplicates
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        lanid_cell = row[lanid_column_index - 1] # Adjust for 0-based index
                # Skip rows with blank LANID
        if not lanid_cell.value:
            continue
        
        ms_name_cell = row[ms_name_column_index - 1] # -1 to convert back to 0-based indexing
        # Handle null or blank values explicitly
        ms_name_value = ms_name_cell.value or ""
        ms_name_value = ms_name_value.lower().strip()

        # Apply cell format to LANID cell if it's a duplicate
        if lanid_cell.value in duplicate_lanids: # lanid_cell is specifically targeted within each row
            lanid_cell.fill = duplicate_fill

def normalize(value):
    """
    Normalizes a value for comparison by converting it to a standard format.
    Parameters:
    value: The value to be normalized, which will be str.
    Returns:
    The normalized value:
    - If the value is NaN (Not a Number), returns None.
    - If the value is a string, strips leading and trailing spaces and converts it to lowercase.
    - If the value is an integer or a float, converts it to a string
    - If the value is a datetime object, converts it to a string in the format 'YYYY-MM-DD'.
    - If the value is of any other type, returns the value unchanged.
    
    This function helps ensure that comparisons between values focus on the content rather than differences in data types or formatting.
    """
    if pd.isna(value):
        return None
    if isinstance(value, str):
        return value.strip().lower()
    if isinstance(value, (int, float)):
        return str(value)  # Convert all numbers to floats for consistent comparison
    if isinstance(value, datetime):
        return value.strftime('%b-%y')  # Convert dates to a standard string format
    return value