import pandas as pd
from config.config_GUI import *
from modules.logger import data_logger
import os
from openpyxl import load_workbook
from identifier_functions_MS.identifier_MS_GUI import *
from modules.formatting import *
from modules.eofy import get_eofy
from modules.get_column_index import get_column_index
from modules.date_extraction import extract_date_from_filename
from modules.skip_column import initiate_skip_column_ms
from modules.missing_employees import identify_missing_employees_ms
from modules.check_mark_fulfilled import check_and_mark_fulfilled_ms
from modules.sanity_checks import sanity_checks_ms
from modules.kill_switch import terminate_process

"""
Global Parameters:
    current_df (DataFrame): DataFrame containing the current month data from the Static Report.
    next_df (DataFrame): DataFrame containing the next month data from the Static Report.
    op_ms_df (DataFrame): DataFrame containing the operational plan data.
    original_op_ms_df (DataFrame): The original DataFrame.
    global_staff_df (DataFrame): Gloval Staff List
    ws (Worksheet): The worksheet to apply the highlights.
    output_directory (str): The directory where the output file will be saved
""" 

def load_data():
    """
    Loads data from the operational plan and static files based on configuration.
    
    Returns:
    tuple: DataFrames for current month, next month, operational plan, global staff list, and original operational plan.
    
    Process:
    1. Define a mapping dictionary for employee group names.
    2. Load current month and next month data from the two Static Report sheets.
    3. Apply employee group mapping and rename columns based on the configuration.
    4. Filter data for the Security domain entries.
    5. Log the count of records loaded from the Static Report.
    6. Load the Op Plan MS data.
    7. Filter out rows with 'MS Resource Type', 'x', missing 'LANID', or 'FTE Resource Type' in the 'Resource Type' column.
    8. Remove duplicates based on 'Employee ID'.
    9. Log the count of records in the Op Plan FTE sheet.
    10. Create a copy of the Op Plan for comparison later.
    11. Load the Global Staff List data.
    12. Filter the Global Staff List to only include employees from the 'Technology' division.
    13. Return the loaded DataFrames.
    If an error occurs during the process, logs the error and returns None for all DataFrames.
    """
    try:
        # Define a mapping dictionary for employee group names
        employee_group_mapping = {
            'CWR - Consultant': 'MS',
            'CWR - Contractor Professional': 'MS',
            'CWR - Contractor Temp Workers': 'MS',
            'CWR - Intern (Work Experience)': 'MS',
            'CWR - Managed Services': 'MS',
            'CWR - Sales Agent': 'MS',
            'CWR - Trainee': 'MS'
        }

        # Load Static Report
        data_logger.info(f"Loading Static Report data from {CONFIG['STATIC_FILE']}...")
        current_df = pd.read_excel(CONFIG['STATIC_FILE'], sheet_name=CONFIG['sheets']['current_month']['sheet_name'],
                                   usecols=CONFIG['sheets']['current_month']['usecols'], header=1)
        next_df = pd.read_excel(CONFIG['STATIC_FILE'], sheet_name=CONFIG['sheets']['next_month']['sheet_name'],
                                usecols=CONFIG['sheets']['next_month']['usecols'], header=1)

        # Pre-formatting datetime, and Employee Category
        for df in [current_df, next_df]:
            df['Employee Group (Name)'] = df['Employee Group (Name)'].replace(employee_group_mapping)
            df.rename(columns=CONFIG['COLUMN_MAPPING_MS'], inplace=True)
        data_logger.info("Date formatting and Name filtering has been applied for Static Report.")

        # Filter for Security domain entries
        current_security_count = len(current_df[(current_df['Domain'] == 'Security') & (current_df['FTE Category'] == 'Non-FTE')])
        next_security_count = len(next_df[(next_df['Domain'] == 'Security') & (next_df['FTE Category'] == 'Non-FTE')])
        
        data_logger.info(f"Security domain: {current_security_count} records from current month, {next_security_count} records from next month.")

        # Load Op Plan
        data_logger.info(f"Loading Op Plan MS data from {CONFIG['OP_FILE']}...")
        op_ms_df = pd.read_excel(CONFIG['OP_FILE'], sheet_name=CONFIG['sheets']['MS']['sheet_name'],
                                 usecols=CONFIG['sheets']['MS']['usecols'], header=3)
        
        # Removing duplicates based on Employee ID
        unique_op_ms_df = op_ms_df.drop_duplicates(subset=['Employee ID'])
        
        op_ms_count = len(unique_op_ms_df[(unique_op_ms_df['Resource Type'] != "MS Resource Type") & 
                                    (unique_op_ms_df['Resource Type'] != "x") &
                                    (unique_op_ms_df['LANID'] != "") &
                                    (unique_op_ms_df['Employee ID'].notna())])

        data_logger.info(f"Security domain: {op_ms_count} unique records from MS sheet.")
        data_logger.info("The figures stated above are estimates, as these also accounted for duplicated entries and cancelled/vacants, etc.")
        
        original_op_ms_df = op_ms_df.copy()

        # Load Global Staff List
        data_logger.info(f"Loading Global Staff List data from {CONFIG['GLOBAL_STAFF_LIST']}...")
        global_staff_df = pd.read_excel(CONFIG['GLOBAL_STAFF_LIST'], sheet_name=CONFIG['sheets']['global']['sheet_name'], 
                                        usecols=CONFIG['sheets']['global']['usecols'], header=0)
        global_staff_df = global_staff_df[global_staff_df['Operational Division (Label)'] == 'Technology']

        return current_df, next_df, op_ms_df, global_staff_df, original_op_ms_df

    except Exception as e:
        data_logger.error(f"Error loading data: {e}")
        return None, None, None, None, None

def merge_data_ms(current_df, next_df, op_ms_df):
    """
    Merges the operational plan data with the current month and next month data
    from the static report. Set datetime object. Initializes the 'Modified' and 'Line Manager' columns.
    
    Returns:
    DataFrame: Merged DataFrame containing all the data with initialized columns.
    
    Process:
    1. Log the start of data processing.
    2. Retrieve and log the current End of Financial Year (EOFY).
    3. Get and log the current date.
    4. Extract and check the file date from the static report filename.
    5. Rename columns in current_df and next_df to match the Op Plan columns based on the provided mapping.
    6. Initialize the 'Modified' column in op_ms_df.
    7. Initialize the 'Line Manager' column in op_ms_df.
    8. Log the completion of data merging and column initialization.
    9. Return the updated op_ms_df.
    """
    data_logger.info("Starting data processing...")
    eofy = get_eofy()
    data_logger.info(f"The current End of Financial Year (EOFY) is: {eofy}")
    
    current_date = datetime.now().strftime('%d-%m-%y')
    data_logger.info(f"The current date is: {current_date}")

    # Extract and check the file date
    file_date = extract_date_from_filename(CONFIG['STATIC_FILE'])
    if not file_date:
        data_logger.error("Failed to extract date from Static Report")
        return None

    # Rename columns in current_df and next_df to match the Op Plan columns based on the mapping
    current_df = current_df.rename(columns=CONFIG['COLUMN_MAPPING_MS'])
    next_df = next_df.rename(columns=CONFIG['COLUMN_MAPPING_MS'])

    # Initialize 'Modified' column
    op_ms_df['Modified'] = ""  # Mark all entries as initially unmodified

    # Initialize 'Line Manager' column
    op_ms_df['Line Manager'] = ""  # Mark all entries as initially unmodified

    data_logger.info("Merging data process completed.")
    return op_ms_df

def highlight_differences(ws, op_ms_df, original_op_ms_df):
    """
    Highlights cells in the modified DataFrame that are different from the original DataFrame.
    Stops checking at the row where multiple columns contain 'x'.
    Skips multiple columns as there are not enough sufficient information for comparison.
    
    Process:
    1. Define the fill pattern for modified cells.
    2. Extract headers and create mappings for original and processed data indices.
    3. Determine the row to stop highlighting based on 'x' count in columns A to G.
    4. Normalize values for comparison.
    5. Identify columns to skip based on predefined skip list.
    6. Iterate over each row in the worksheet:
       a. Skip rows based on specific conditions (e.g., 'Vacant', 'Role handed back').
       b. For matching rows in original and processed data, compare cell values.
       c. Highlight cells with differences using the defined fill pattern.
    """
    modified_fill = PatternFill(start_color="7EC8E3", end_color="7EC8E3", fill_type="solid") # Light Blue for sanity check

    headers = {cell.value: cell.column_letter for cell in ws[1]}
    original_index_map = original_op_ms_df.groupby(['Employee ID', 'Resource Name', 'Start Date']).apply(lambda x: x.index.tolist()).to_dict()
    processed_index_map = op_ms_df.groupby(['Employee ID', 'Resource Name', 'Start Date']).apply(lambda x: x.index.tolist()).to_dict()

    stop_highlighting_row = None
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        x_count = sum(cell.value == 'x' for cell in row)
        if x_count > 1:
            stop_highlighting_row = row[0].row
            break

    skip_columns_list = [
        'MS Daily Rate', 'Annualised FY23 Cost or (Stretch) $', 'Squad', 'Service', 
        'Asset', 'Product', 'Tech Financial Reform', 'T&M / FP', 'Comments',
        'Free Input', 'Run %', 'Divisional Change %', 'Tech Projects %', 'Total %'
    ]

    # Identify columns to skip based on header names
    skip_columns = [col_letter for col_name, col_letter in headers.items() if col_name in skip_columns_list]

    for row in ws.iter_rows(min_row=2, max_row=(stop_highlighting_row or ws.max_row) - 1):
        row_idx = row[0].row - 2
        resource_name = op_ms_df.at[row_idx, 'Resource Name']
        employee_id = op_ms_df.at[row_idx, 'Employee ID']
        lanid = op_ms_df.at[row_idx, 'LANID']
        # tech_area = op_ms_df.at[row_idx, 'Tech Area']
        start_date = op_ms_df.at[row_idx, 'Start Date']

        if pd.isna(employee_id) or pd.isna(lanid):
            continue
        
        if (employee_id, resource_name, start_date) in original_index_map:
            original_row_indices = original_index_map[(employee_id, resource_name, start_date)]
            processed_row_indices = processed_index_map[(employee_id, resource_name, start_date)]
            for original_row_idx in original_row_indices:
                for processed_row_idx in processed_row_indices:
                    for col_name, col_letter in headers.items():
                        if col_name in original_op_ms_df.columns and col_name in op_ms_df.columns:
                            if col_letter in skip_columns:
                                continue
                            
                            original_value = original_op_ms_df.at[original_row_idx, col_name]
                            modified_value = op_ms_df.at[processed_row_idx, col_name]
                            
                            if pd.isna(modified_value) or pd.isna(original_value):
                                continue

                            original_value = normalize(original_value)
                            modified_value = normalize(modified_value)

                            if original_value != modified_value:
                                cell = row[ws[col_letter + str(row[0].row)].col_idx - 1]
                                cell.fill = modified_fill

                                # Debug print to see what's being compared
                                print(f"Highlighting cell {cell.coordinate} - Original: {original_value}, Modified: {modified_value}")

def save_data(op_ms_df, original_op_ms_df, output_directory):
    """
    Saves the processed data to an Excel file with specific formatting applied to rows based on their role status.
    Includes rows with specific keywords from the original dataset.
    
    Returns:
    str: The path to the saved Excel file.
    
    Process:
    1. Generate a timestamp and create the output file path.
    2. Check if the output file already exists and log if it will be overwritten.
    3. Filter out rows where "Resource Type" is "FTE Resource Type".
    4. Save the filtered DataFrame to an Excel file.
    5. Load the workbook and worksheet from the saved file.
    6. Get column indices for 'LANID', 'FTE Name', 'Start Date', and 'End Date'.
    7. Format duplicated 'LANID' cells.
    8. Apply date formatting to 'Start Date' and 'End Date' columns.
    9. Highlight differences between the processed and original DataFrames.
    10. Highlight vacant and stretch roles.
    11. Save the formatted workbook.
    12. Log the save operation and return the output file path.
    """
    timestamp = datetime.now().strftime("%Y-%m-%d")
    output_file = os.path.join(output_directory, f"Security (MS) - {timestamp}.xlsx")

    # Check if the output is existed
    if os.path.exists(output_file):
        data_logger.info(f"File already exists. Overwriting... {output_file}")

    # Filter out rows where "Resource Type" = "MS Resource Type" and create 'MS' sheet
    filtered_df = op_ms_df[(op_ms_df['Resource Type'] != "MS Resource Type")]
    filtered_df.to_excel(output_file, index=False, sheet_name='MS', engine='openpyxl')

    # Load workbook and worksheet
    wb = load_workbook(output_file)
    ws = wb.active

    lanid_column_index = get_column_index(ws, 'LANID')
    ms_name_column_index = get_column_index(ws, 'Resource Name')
    format_duplicate_lanid_ms(ws, lanid_column_index, ms_name_column_index)

    start_date_index = get_column_index(ws, 'Start Date')
    end_date_index = get_column_index(ws, 'End Date')

    # Apply comprehensive formatting to the worksheet
    apply_date_format(wb, ws, start_date_index, end_date_index)

    # Highlight modified cells
    highlight_differences(ws, op_ms_df, original_op_ms_df)

    wb.save(output_file)
    data_logger.info(f"Data saved to {output_file}")

    return output_file

def process_ms(op_plan_path, static_report_path, global_staff_path, output_directory, result):
    """
    Executes the MS Op Plan Automation Script.
    
    Parameters:
    op_plan_path (str): The file path to the operational plan.
    static_report_path (str): The file path to the static report.
    global_staff_path (str): The file path to the global staff list.
    output_directory (str): The directory where the output file will be saved.
    result: Required if you are running this from the GUI, if you want to run it directly from fte_GUI.py, 
    then remove this parameter and provide the path to file at the end of the script.
    
    Process:
    1. Set the file paths for the operational plan and static report.
    2. Log the start of the process and the current timestamp.
    3. Load data from the specified files.
       - If data loading fails, log an error and terminate the script.
    4. Check for termination signal after each major step.
    5. Identify missing employees from the original operational plan.
    6. Mark fulfilled rows.
    7. Initialize the 'Skip' column.
    8. Perform sanity checks on the data.
    9. Extract the date from the static report filename for consistent use.
       - If date extraction fails, log an error and terminate the script.
    10. Merge the operational plan data with current and next month data.
        - If merging fails, log an error and terminate the script.
    11. Process data through various scenario functions to identify specific changes:
        - Exits, new joiners, transfers in/out, grade changes, internal mobility, conversions, line manager changes, location changes.
    12. Save the processed data to an Excel file.
    13. Highlight differences and vacant/stretch roles in the saved workbook.
    14. Save the workbook after highlighting differences.
    15. Log the completion time of the script and the duration of the execution.
    
    Exceptions:
    - Logs any unexpected errors and terminates the process gracefully.
    """
    
    set_op_plan_path(op_plan_path)
    set_static_report_path(static_report_path)
    set_global_staff_list_path(global_staff_path)

    try:
        # Start Script
        data_logger.info('-----------------------------------------------------------------------------------------------------------------------------------')
        data_logger.info("Executing MS Op Plan Automation Script ...")
        start_time = datetime.now()

        # Load Data
        current_df, next_df, op_ms_df, global_staff_df, original_op_ms_df = load_data()
        if op_ms_df is None:
            data_logger.error("Failed to load Op Plan MS data. Exiting script")
            return

        # Check for termination
        if terminate_process.is_set():
            data_logger.info("Process terminated by user.")
            return

        # Initiate Skip column
        op_ms_df = initiate_skip_column_ms(op_ms_df)

        # Identify Missing Employees from the original Op Plan
        op_ms_df = identify_missing_employees_ms(current_df, next_df, op_ms_df)

        # Check for termination
        if terminate_process.is_set():
            data_logger.info("Process terminated by user.")
            return

        # Check and mark entries as fulfilled
        op_ms_df = check_and_mark_fulfilled_ms(current_df, next_df, op_ms_df)

        # Sanity Checks
        op_ms_df = sanity_checks_ms(current_df, next_df, op_ms_df)

        # Extract the date once and use it throughout to avoid repeated extraction
        file_date = extract_date_from_filename(CONFIG['STATIC_FILE'])
        if not file_date:
            data_logger.error("Failed to extract date from Static Report. Exiting script")
            return
            
        # Process data (initial processing)
        op_ms_df = merge_data_ms(current_df, next_df, op_ms_df)
        if op_ms_df is None:
            data_logger.error("Merging data process failed. Exiting script")
            return

        # Process data
        scenario_functions_ms = [
            identify_exits_ms,
            identify_new_joiners_ms,
            identify_transfers_in_ms,
            identify_transfers_out_ms,
            identify_conversions_fte_to_cwr,
            identify_conversions_cwr_to_fte,
            identify_internal_mobility_ms,
            identify_line_manager_changes_ms,
            identify_location_changes_ms,
        ]

        for func in scenario_functions_ms:
            # Check for termination
            if terminate_process.is_set():
                data_logger.info("Process terminated by user.")
                return
            op_ms_df = func(current_df, next_df, op_ms_df, file_date)
        
        # Save Data
        output_file = save_data(op_ms_df, original_op_ms_df, output_directory)

        # Highlight differences
        wb = load_workbook(output_file)
        ws = wb.active
        highlight_differences(ws, op_ms_df, original_op_ms_df)

        # Save the workbook after highlighting differences
        wb.save(output_file)

        # Log script completion time
        end_time = datetime.now()
        duration = (end_time - start_time).total_seconds()
        data_logger.info(f"Script execution completed successfully in {duration} seconds.")
        
    except Exception as e:
        data_logger.error(f"An unexpected error occurred, terminating process: {e}.", exc_info=True)

"""
Provide the path to the file you want to run the script directly from the file.

# if __name__ == "__main__":
#     test_op_plan_path = "My Documents\\An\\VS Learning\\Practice\\New-Test files\\Copy of Op - May.xlsx"
#     test_static_report_path = "My Documents\\An\VS Learning\\Practice\\New-Test files\\Copy of Static 240630.xlsx"
#     test_output_directory = "My Documents\\An\\VS Learning\\Practice\\New-Test files"
#     test_global_staff_list = "My Documents\\An\\VS Learning\\Practice\\New-Test files\\Copy of Staff list - May.xlsx"
#     process_ms(test_op_plan_path, test_static_report_path, test_global_staff_list, test_output_directory)
"""